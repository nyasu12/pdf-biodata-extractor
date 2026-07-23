import datetime
import os
import re
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def normalize_date_string(value):
    text = str(value or "").strip()
    if not text:
        return ""

    text = re.sub(r"\s+", " ", text)
    text = re.sub(
        r"\b([A-Za-z]{3,9})\.(\s+\d{1,2},\s+\d{4}\b)",
        r"\1\2",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\bSEPT\b", "SEP", text, flags=re.IGNORECASE)

    patterns = [
        "%B %d, %Y",
        "%b %d, %Y",
        "%B %d %Y",
        "%b %d %Y",
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
    ]

    for fmt in patterns:
        try:
            dt = datetime.datetime.strptime(text.title(), fmt)
            return dt.strftime("%Y/%m/%d")
        except ValueError:
            pass

    return text


def parse_date_for_sort(value):
    text = normalize_date_string(value)
    if not text:
        return None

    try:
        return datetime.datetime.strptime(text, "%Y/%m/%d")
    except ValueError:
        return None


def _employment_key(profile: Dict[str, Any]) -> str:
    return (
        profile.get("extraction", {})
        .get("employment", {})
        .get("key", "employment_history")
    )


def _employment_records(row: Dict[str, Any], profile: Dict[str, Any]) -> List[Dict[str, Any]]:
    value = row.get(_employment_key(profile), [])
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _record_matches(record: Dict[str, Any], filter_config: Optional[Dict[str, Any]]) -> bool:
    if not filter_config:
        return True

    country = str(record.get("country", "") or "").strip().casefold()

    if "country" in filter_config:
        expected = str(filter_config.get("country", "") or "").strip().casefold()
        if country != expected:
            return False

    if "not_country" in filter_config:
        blocked = str(filter_config.get("not_country", "") or "").strip().casefold()
        if country == blocked:
            return False

    return True


def _filter_records(records: Iterable[Dict[str, Any]], filter_config: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [record for record in records if _record_matches(record, filter_config)]


def _sort_records(records: List[Dict[str, Any]], sort_by: Optional[str]) -> List[Dict[str, Any]]:
    if not sort_by:
        return list(records)

    parseable = []
    unparseable = []
    for index, record in enumerate(records):
        parsed = parse_date_for_sort(record.get(sort_by, ""))
        if parsed is None:
            unparseable.append((index, record))
        else:
            parseable.append((parsed, index, record))

    parseable.sort(key=lambda item: (item[0], item[1]))
    return [item[2] for item in parseable] + [item[1] for item in unparseable]


def _records_for_country(row: Dict[str, Any], profile: Dict[str, Any], country: str) -> List[Dict[str, Any]]:
    records = _employment_records(row, profile)
    return _filter_records(records, {"country": country})


def _latest_record(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not records:
        return {}

    dated = []
    for index, record in enumerate(records):
        parsed = parse_date_for_sort(record.get("start_date", ""))
        if parsed is not None:
            dated.append((parsed, index, record))

    if dated:
        dated.sort(key=lambda item: (item[0], item[1]))
        return dated[-1][2]
    return records[-1]


def _format_value(value: Any, format_name: Optional[str]) -> Any:
    if format_name == "date":
        return normalize_date_string(value)
    if value is None:
        return ""
    return value


def _resolve_column_value(row: Dict[str, Any], column: Dict[str, Any], profile: Dict[str, Any]) -> Any:
    source = column.get("source", "")

    if source == "employment_count":
        country = str(column.get("country", ""))
        return len(_records_for_country(row, profile, country))

    if source in {"employment_latest_start", "employment_latest_end"}:
        country = str(column.get("country", ""))
        latest = _latest_record(_records_for_country(row, profile, country))
        field = "start_date" if source.endswith("start") else "end_date"
        return _format_value(latest.get(field, ""), column.get("format"))

    return _format_value(row.get(source, ""), column.get("format"))


def _build_repeat_metadata(data_list: List[Dict[str, Any]], profile: Dict[str, Any]):
    repeat_sections = profile.get("excel", {}).get("repeat_sections", []) or []
    metadata = []

    for section in repeat_sections:
        max_records = 0
        filter_config = section.get("filter")
        sort_by = section.get("sort_by")
        for row in data_list:
            records = _filter_records(_employment_records(row, profile), filter_config)
            records = _sort_records(records, sort_by)
            max_records = max(max_records, len(records))

        metadata.append((section, max_records))

    return metadata


def save_to_excel(data_list, output_path, profile):
    output_path = str(output_path)
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    excel_config = profile.get("excel", {}) or {}
    base_columns = excel_config.get("columns", []) or []
    repeat_metadata = _build_repeat_metadata(data_list, profile)

    headers: List[str] = []
    widths: List[float] = []

    for column in base_columns:
        headers.append(str(column.get("header", column.get("source", ""))))
        widths.append(float(column.get("width", 20)))

    for section, max_records in repeat_metadata:
        fields = section.get("fields", []) or []
        for index in range(1, max_records + 1):
            for field in fields:
                headers.append(str(field.get("header", "{index}")).format(index=index))
                widths.append(float(field.get("width", 20)))

    rows = []
    for row in data_list:
        out: Dict[str, Any] = {}

        for column in base_columns:
            header = str(column.get("header", column.get("source", "")))
            out[header] = _resolve_column_value(row, column, profile)

        for section, max_records in repeat_metadata:
            records = _filter_records(_employment_records(row, profile), section.get("filter"))
            records = _sort_records(records, section.get("sort_by"))
            fields = section.get("fields", []) or []

            for index in range(1, max_records + 1):
                record = records[index - 1] if index <= len(records) else {}
                for field in fields:
                    header = str(field.get("header", "{index}")).format(index=index)
                    source = field.get("source", "")
                    out[header] = _format_value(record.get(source, ""), field.get("format"))

        rows.append(out)

    df = pd.DataFrame(rows, columns=headers)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = str(excel_config.get("sheet_name", "Extracted Data"))[:31]

    freeze_panes = excel_config.get("freeze_panes")
    if freeze_panes:
        ws.freeze_panes = str(freeze_panes)

    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    row_height = float(excel_config.get("row_height", 25))
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(output_path)
